import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time
from datetime import datetime
from ttkthemes import ThemedTk
import os




def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Create a class for the GUI application
class SmartrLogistics:
    # Define the attributes and widgets of the application
    def __init__(self):
        self.file_path=None
        # Create the root window
        self.root = ThemedTk()
        self.root.title("Smartr Logistics")
        # Create the frame
        self.frame = ttk.Frame(self.root)
        self.frame.pack(fill=tk.BOTH, expand=True)
        # Create the widgets frame
        self.widgets_frame = ttk.LabelFrame(self.frame)
        self.widgets_frame.grid(row=0, column=0, padx=20, pady=10)
        self.name_entry = ttk.Entry(self.widgets_frame)
        self.name_entry.insert(0, "User Name")
        self.name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")
        self.password_entry = ttk.Entry(self.widgets_frame)
        self.password_entry.insert(0, "Password")
        self.password_entry.bind("<FocusIn>", lambda e: self.password_entry.delete('0', 'end'))
        self.password_entry.grid(row=3, column=0, padx=5, pady=(0, 5), sticky="ew")
        # Create the buttons
        self.button = ttk.Button(self.widgets_frame, text="Click to update", command=self.update_data)
        self.button.grid(row=10, column=0, padx=5, pady=5, sticky="nsew")
        self.upload_button = tk.Button(self.root, text="Upload File", command=self.upload_file)
        self.upload_button.pack(pady=20)
        self.refresh_button = ttk.Button(self.widgets_frame, text="Refresh File", command=self.refresh_data)
        self.refresh_button.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")
        self.download_failed_button = tk.Button(self.root, text="Download FailedAWBs", command=self.download_failed_awbs)
        self.download_failed_button.pack(pady=10)

        self.download_passed_button = tk.Button(self.root, text="Download PassedAWBs", command=self.download_passed_awbs)
        self.download_passed_button.pack(pady=10)

        self.download_failed_button = tk.Button(self.root, text="Backup_file", command=self.Backup_file)
        self.download_failed_button.pack(pady=10)
        # Create the notebook widget
        self.notebook = ttk.Notebook(self.frame)
        self.notebook.grid(row=0, column=1, columnspan=2, pady=10)
        # Create the first treeview for the passed data
        self.treeFrame1 = ttk.Frame(self.notebook)
        self.treeScroll1 = ttk.Scrollbar(self.treeFrame1)
        self.treeScroll1.pack(side="right", fill="y")
        self.cols1 = ("AWB Number", "Date", "Route_ID", "Receiver_Name")
        self.treeview1 = ttk.Treeview(self.treeFrame1, show="headings",
                            yscrollcommand=self.treeScroll1.set, columns=self.cols1 ,height=15)
        self.treeview1.column("AWB Number", width=100)
        self.treeview1.column("Date", width=100)
        self.treeview1.column("Route_ID", width=100)
        self.treeview1.column("Receiver_Name", width=800)
        self.treeview1.pack()
        self.treeScroll1.config(command=self.treeview1.xview)
        # Add the first treeview to the notebook as a tab
        self.notebook.add(self.treeFrame1, text="Passed Data")
        self.passed_awbs_data()
        # Create the second treeview for the failed data
        self.treeFrame2 = ttk.Frame(self.notebook)
        self.treeScroll2 = ttk.Scrollbar(self.treeFrame2)
        self.treeScroll2.pack(side="right", fill="y")
        self.cols2 = ("AWB Number", "Date","Route_ID","Receiver_Name","Update Status")
        self.treeview2 = ttk.Treeview(self.treeFrame2, show="headings",
                            yscrollcommand=self.treeScroll2.set, columns=self.cols2, height=15)
        self.treeview2.column("AWB Number", width=100)
        self.treeview2.column("Date", width=100)
        self.treeview2.column("Route_ID", width=100)
        self.treeview2.column("Receiver_Name", width=100)
        self.treeview2.column("Update Status", width=700)
        self.treeview2.pack()
        self.treeScroll2.config(command=self.treeview2.xview)
        self.notebook.add(self.treeFrame2, text="Failed Data")
        self.failed_awbs_data()
        

    def upload_file(self):
        global file_path
        # Add your file upload logic here
        # Load the first workbook
        workbook1 = openpyxl.load_workbook(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\PassedAWBs.xlsx"))
        # Select the sheet by name
        sheet1 = workbook1.active
        # Clear all data in the sheet
        sheet1.delete_rows(2, sheet1.max_row)
        # Delete the data from the treeview
        for item in self.treeview1.get_children():
            self.treeview1.delete(item)
        # Save the changes
        workbook1.save(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\PassedAWBs.xlsx"))

        # Load the second workbook
        workbook2 = openpyxl.load_workbook(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\FailedAWBs.xlsx"))
        # Select the sheet by name
        sheet2 = workbook2.active
        # Clear all data in the sheet  
        sheet2.delete_rows(2, sheet2.max_row)
        # Delete the data from the treeview
        for item in self.treeview2.get_children():
            self.treeview2.delete(item)
        # Save the changes
        workbook2.save(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\FailedAWBs.xlsx"))

        # Ask user to select a file
        file_path = filedialog.askopenfilename(title="Select a file")
        return self.file_path

    def passed_awbs_data(self):
        path = "C:\\Users\\Admin\\Desktop\\WebScraping\\PassedAWBs.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        list_values = list(sheet.values)
        for col_name in list_values[0]:
            self.treeview1.heading(col_name, text=col_name)

        for value_tuple in list_values[1:]:
            self.treeview1.insert('', tk.END, values=value_tuple)

    def failed_awbs_data(self):
        path = "C:\\Users\\Admin\\Desktop\\WebScraping\\FailedAWBs.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        list_values = list(sheet.values)
        for col_name in list_values[0]:
            self.treeview2.heading(col_name, text=col_name)

        for value_tuple in list_values[1:]:
            self.treeview2.insert('', tk.END, values=value_tuple)

    def refresh_data(self):
        for item in self.treeview1.get_children():
            self.treeview1.delete(item)
        self.passed_awbs_data()
        for item in self.treeview2.get_children():
            self.treeview2.delete(item)
        self.failed_awbs_data()

    def download_failed_awbs(self):
        self.download_file(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\FailedAWBs.xlsx"))

    def download_passed_awbs(self):
        self.download_file(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\PassedAWBs.xlsx"))

    def download_file(self, filename):
        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title=f"Save {filename}")
            if file_path:
                workbook = openpyxl.load_workbook(filename)
                workbook.save(file_path)
                print(f"{filename} downloaded successfully.")
            else:
                print(f"{filename} download canceled.")
        except Exception as e:
            print(f"Error downloading {filename}: {e}")

    def Backup_file(self):
        # copying data from passedAWBs sheet to MasterPODSheet sheet
        master_wb = openpyxl.load_workbook(resource_path("C:\\Users\\Admin\\Documents\\Backup_of_pod.xlsx"))
        successful_wb = openpyxl.load_workbook(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\PassedAWBs.xlsx"))
        master_sheet = master_wb.active
        successful_sheet = successful_wb.active

        for row in successful_sheet.iter_rows():
            row_values=[cell.value for cell in row]
            master_sheet.append(row_values)
            
        master_wb.save(resource_path("C:\\Users\\Admin\\Documents\\Backup_of_pod.xlsx"))
        print("File Backuped Successfully")
        master_wb.close()
        successful_wb.close()

        

    def update_data(self):
    # Access the instance variable file_path
        global file_path
        user_name = self.name_entry.get()
        password = self.password_entry.get()

        if file_path:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            #  # Set the path to your chromedriver.exe
            # try:
            #     # Try to execute this block
            #     service = ChromeService(executable_path="chromedriver.exe")
            #     print("Old version of chromedriver is being used")
               
            # except Exception as e:
            #     # If there's an error, execute this block
            #     print(f"Error occurred: {e}")
            #     service = ChromeService(ChromeDriverManager().install())
            #     print("latest version installed successfully")

            # chrome_options = Options()
            # chrome_options.add_argument("--headless")

            # # Initialize the WebDriver with the specified options
            # driver = webdriver.Chrome(service=service,options=chrome_options)

            service = ChromeService(executable_path="chromedriver.exe")
            chrome_options = Options()
            chrome_options.add_argument("--headless")

            # Initialize the WebDriver with the specified options
            driver = webdriver.Chrome(service=service,options=chrome_options)

            try:
                driver.get("https://login.smartr.in")

                # Wait for 5 seconds until the browser finds the element id
                WebDriverWait(driver, 5).until(
                    EC.presence_of_all_elements_located((By.ID, "ctl00_ContentPlaceHolder1_txtUserName"))
                )
                input_element = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtUserName")
                input_element.clear()
                input_element.send_keys(user_name + Keys.TAB)

                WebDriverWait(driver,5).until(
                    EC.presence_of_all_elements_located((By.ID,"ctl00_ContentPlaceHolder1_txtPwd"))
                    )

                input_element = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtPwd")
                input_element.clear()
                input_element.send_keys(password + Keys.ENTER)

                # Searching for the revoke AWB status page
                WebDriverWait(driver, 5).until(
                    EC.presence_of_all_elements_located((By.ID, "ctl00_txtSearchLink"))
                )

                input_element = driver.find_element(By.ID, "ctl00_txtSearchLink")
                text = "Revoke AWB Status"

                # Slowing down the speed of send key in the element
                for character in text:
                    input_element.send_keys(character)
                    time.sleep(0.4)

                # After entering all the characters in the element, press the enter button
                input_element.send_keys(Keys.ENTER)

                WebDriverWait(driver, 5).until(
                    EC.presence_of_all_elements_located((By.ID, "__tab_ctl00_ContentPlaceHolder1_TabContainer1_TabPanel3"))
                )

                # Clicking the Revoke element
                input_element = driver.find_element(By.ID, "__tab_ctl00_ContentPlaceHolder1_TabContainer1_TabPanel3")
                input_element.click()

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) == 4:
                        awb_number, date, route_id, receiver_name = row
                     # rest of your code..
                    else:
                        # Handle the case where the length of row is not 4
                        print("Invalid row format:", row)
                    date_string = date.strftime("%Y-%m-%d")
                    date_object = datetime.strptime(date_string, "%Y-%m-%d")
                    awb_date = date_object.strftime("%d/%m/%Y")


                    # Sending the input to the AWB element
                    input_element = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_TabContainer1_TabPanel3_txtDlvAWBNumber")
                    input_element.clear()
                    input_element.send_keys(awb_number)

                    # This section can be modified based upon the data duplication
                    # Sending the input to the Date element
                    input_element = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_TabContainer1_TabPanel3_textIssueDate_txtDate")
                    input_element.clear()
                    input_element.send_keys(awb_date)

                    # Sending the input to the RouterID element
                    input_element = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_TabContainer1_TabPanel3_txtRouteID")
                    input_element.clear()
                    input_element.send_keys(route_id)

                    # Sending the input to the ReceiverName element
                    input_element = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_TabContainer1_TabPanel3_txtReceiverName")
                    input_element.clear()
                    input_element.send_keys(receiver_name)
                    # Last section of modified element for faster data pass to the element

                    # Clicking the Enter button at the end
                    form_element = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_TabContainer1_TabPanel3_btnUpdateDelivered")
                    form_element.send_keys(Keys.ENTER)

                    # Reading the AWB update status
                    try:
                        WebDriverWait(driver, 5).until(
                            EC.presence_of_all_elements_located((By.ID, "ctl00_ContentPlaceHolder1_lblStatus"))
                        )
                        awb_status = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_lblStatus")
                        value = awb_status.text
                        if value != "AWB updated successfully.":

                            # Open the FailedAWBs.xlsx file
                            failed_workbook = openpyxl.load_workbook(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\FailedAWBs.xlsx"))
                            failed_sheet = failed_workbook.active
                            row_value = [awb_number, date, route_id, receiver_name, value]

                            # Append failed AWB to FailedAWBs.xlsx
                            failed_sheet.append(row_value)

                            # Save the changes to FailedAwbs.xlsx
                            failed_workbook.save(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\FailedAWBs.xlsx"))

                            print(awb_number, date, route_id, receiver_name, ": ", value)
                            self.treeview2.insert('', tk.END, values=row_value)
                        else:
                            # Open the PassedAWBs.xlsx file
                            successful_wb = openpyxl.load_workbook(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\PassedAWBs.xlsx"))
                            successful_sheet = successful_wb.active
                            row_value = [awb_number, date, route_id, receiver_name]

                            # Append successful AWB to PassedAWBs.xlsx
                            successful_sheet.append(row_value)

                            # Save the changes to Successful.xlsx
                            successful_wb.save(resource_path("C:\\Users\\Admin\\Desktop\\WebScraping\\PassedAWBs.xlsx"))

                            print("else block code: ", awb_number, date, route_id, receiver_name, ": ", value)
                            self.treeview1.insert('', tk.END, values=row_value)

                    except (TimeoutException, NoSuchElementException):
                        # Handle the case where the element is not found within the timeout
                        print("Element with ID '#")
                    
                    

            finally:
                driver.quit()
                workbook.close()


# Create an instance of the SmartrLogistics class
app = SmartrLogistics()

# Run the Tkinter event loop
app.root.mainloop()
